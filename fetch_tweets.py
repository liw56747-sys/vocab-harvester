"""
Twitter 推文抓取工具

通过 twikit 库从用户主页 URL 抓取推文，导出为 XLSX 和 CSV 文件。

使用方法:
    python fetch_tweets.py

首次运行需要登录 Twitter 账号，之后会保存 cookie 免登录。
"""

import asyncio
import csv
import os
import re
import sys
from datetime import datetime
from pathlib import Path

# ── 配置 ──────────────────────────────────────────────────

COOKIE_FILE = Path(__file__).parent / ".twitter_cookies.json"
OUTPUT_DIR = Path(__file__).parent / "data"
DEFAULT_TWEET_COUNT = 50  # 每个用户默认抓取推文数

# 从 URL 提取用户名
_URL_RE = re.compile(r"(?:https?://)?(?:www\.|mobile\.)?(?:twitter\.com|x\.com)/(\w+)")


def extract_username(text: str) -> str | None:
    """从 URL 或文本中提取 Twitter 用户名"""
    text = text.strip().rstrip("/")
    # 直接就是用户名
    if re.match(r"^\w+$", text):
        return text
    # URL 格式
    m = _URL_RE.match(text)
    if m:
        name = m.group(1)
        # 排除 status/tweets 等路径
        if name.lower() not in ("status", "i", "search", "home", "explore", "settings"):
            return name
    return None


# ── 登录 ──────────────────────────────────────────────────

async def login(client):
    """登录 Twitter，优先使用 cookie，失败则交互式登录"""
    from twikit import Client

    # 尝试加载 cookie
    if COOKIE_FILE.exists():
        print("[*] 发现已保存的登录信息，尝试免登录...")
        try:
            client.load_cookies(str(COOKIE_FILE))
            return True
        except Exception as e:
            print(f"[!] Cookie 加载失败: {e}")

    # 交互式登录
    print("\n=== Twitter 登录 ===")
    print("（首次使用需要登录，之后会自动免登录）\n")

    username = input("Twitter 用户名: ").strip()
    email = input("邮箱: ").strip()
    password = input("密码: ").strip()

    if not all([username, email, password]):
        print("[!] 登录信息不完整")
        return False

    try:
        print("[*] 登录中...")
        await client.login(
            auth_info_1=username,
            auth_info_2=email,
            password=password,
        )
        # 保存 cookie
        client.save_cookies(str(COOKIE_FILE))
        print(f"[+] 登录成功，Cookie 已保存至 {COOKIE_FILE.name}")
        return True
    except Exception as e:
        print(f"[!] 登录失败: {e}")
        return False


# ── 抓取 ──────────────────────────────────────────────────

async def fetch_user_tweets(client, username: str, count: int = DEFAULT_TWEET_COUNT) -> list[dict]:
    """抓取指定用户的推文"""
    print(f"\n[*] 正在获取 @{username} 的推文 (目标 {count} 条)...")

    try:
        user = await client.get_user_by_screen_name(username)
    except Exception as e:
        print(f"[!] 用户 @{username} 不存在或获取失败: {e}")
        return []

    tweets_data = []
    fetched = 0

    try:
        tweets = await user.get_tweets("Tweets", count=min(count, 20))

        while tweets and fetched < count:
            for tweet in tweets:
                if fetched >= count:
                    break

                tweet_url = f"https://x.com/{username}/status/{tweet.id}"

                tweets_data.append({
                    "tweet_id": str(tweet.id),
                    "author": f"@{username}",
                    "author_name": getattr(user, "name", username),
                    "content": tweet.text or "",
                    "created_at": str(tweet.created_at) if tweet.created_at else "",
                    "likes": getattr(tweet, "favorite_count", 0) or 0,
                    "retweets": getattr(tweet, "retweet_count", 0) or 0,
                    "replies": getattr(tweet, "reply_count", 0) or 0,
                    "quotes": getattr(tweet, "quote_count", 0) or 0,
                    "url": tweet_url,
                    "is_retweet": getattr(tweet, "retweeted_tweet", None) is not None,
                    "is_reply": getattr(tweet, "in_reply_to", None) is not None,
                })
                fetched += 1

            # 翻页
            if fetched < count:
                try:
                    tweets = await tweets.next()
                except Exception:
                    break

    except Exception as e:
        print(f"[!] 抓取过程中出错: {e}")

    print(f"[+] @{username}: 成功抓取 {len(tweets_data)} 条推文")
    return tweets_data


# ── 导出 ──────────────────────────────────────────────────

def export_csv(tweets: list[dict], filepath: Path):
    """导出为 CSV"""
    if not tweets:
        return

    fieldnames = [
        "tweet_id", "author", "author_name", "content", "created_at",
        "likes", "retweets", "replies", "quotes", "url",
        "is_retweet", "is_reply",
    ]

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(tweets)

    print(f"[+] CSV 已保存: {filepath}")


def export_xlsx(tweets: list[dict], filepath: Path):
    """导出为 XLSX"""
    if not tweets:
        return

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "推文数据"

    # 表头
    headers = [
        ("推文ID", 18),
        ("作者", 16),
        ("作者昵称", 16),
        ("推文内容", 60),
        ("发布时间", 22),
        ("点赞", 8),
        ("转发", 8),
        ("回复", 8),
        ("引用", 8),
        ("链接", 45),
        ("是否转发", 10),
        ("是否回复", 10),
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = width

    # 修正列宽（超过 Z 列的情况用 openpyxl 内置方法）
    from openpyxl.utils import get_column_letter
    for col, (_, width) in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 数据行
    content_font = Font(name="Microsoft YaHei", size=10)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    normal_align = Alignment(vertical="top")

    keys = [
        "tweet_id", "author", "author_name", "content", "created_at",
        "likes", "retweets", "replies", "quotes", "url",
        "is_retweet", "is_reply",
    ]

    for row_idx, tweet in enumerate(tweets, 2):
        for col_idx, key in enumerate(keys, 1):
            val = tweet.get(key, "")
            if key in ("is_retweet", "is_reply"):
                val = "是" if val else "否"
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = content_font
            cell.border = thin_border
            if key == "content":
                cell.alignment = wrap_align
            elif key in ("likes", "retweets", "replies", "quotes"):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = normal_align

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
    print(f"[+] XLSX 已保存: {filepath}")


# ── 主流程 ────────────────────────────────────────────────

async def main():
    from twikit import Client

    print("=" * 50)
    print("  Twitter 推文抓取工具")
    print("=" * 50)

    client = Client("en-US")

    # 登录
    if not await login(client):
        print("[!] 登录失败，退出")
        return

    # 输入用户
    print("\n--- 输入目标用户 ---")
    print("支持格式: 用户名 / 主页URL（每行一个，输入空行结束）\n")

    usernames = []
    while True:
        line = input("> ").strip()
        if not line:
            break
        name = extract_username(line)
        if name and name not in usernames:
            usernames.append(name)
            print(f"  + 已添加: @{name}")
        elif name in usernames:
            print(f"  ~ 已存在: @{name}")
        else:
            print(f"  ! 无法识别: {line}")

    if not usernames:
        print("[!] 未输入任何用户")
        return

    # 抓取数量
    count_input = input(f"\n每用户抓取条数 (默认 {DEFAULT_TWEET_COUNT}): ").strip()
    count = int(count_input) if count_input.isdigit() and int(count_input) > 0 else DEFAULT_TWEET_COUNT

    # 开始抓取
    print(f"\n{'=' * 50}")
    print(f"  开始抓取 {len(usernames)} 个用户，每人最多 {count} 条")
    print(f"{'=' * 50}")

    all_tweets = []
    for username in usernames:
        tweets = await fetch_user_tweets(client, username, count)
        all_tweets.extend(tweets)
        # 请求间隔，避免被限流
        if username != usernames[-1]:
            await asyncio.sleep(2)

    if not all_tweets:
        print("\n[!] 未抓取到任何推文")
        return

    # 导出
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"tweets_{timestamp}"

    csv_path = OUTPUT_DIR / f"{base_name}.csv"
    xlsx_path = OUTPUT_DIR / f"{base_name}.xlsx"

    print(f"\n{'=' * 50}")
    print(f"  共抓取 {len(all_tweets)} 条推文，正在导出...")
    print(f"{'=' * 50}")

    export_csv(all_tweets, csv_path)
    export_xlsx(all_tweets, xlsx_path)

    print(f"\n[+] 完成！文件保存在:")
    print(f"    CSV:  {csv_path}")
    print(f"    XLSX: {xlsx_path}")
    print(f"\n提示: 可将 CSV/XLSX 中的数据导入 vocab-harvester 进行黑词分析")


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n\n[*] 已中断")
    except Exception as e:
        print(f"\n[!] 程序异常: {e}")
        import traceback
        traceback.print_exc()
