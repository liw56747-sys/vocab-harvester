"""v1.6.6: 定时任务结果落盘测试（CSV / xlsx）；v1.7.3: 评论单独成行 + type/parent_id 列"""

import csv
import os

import pytest
from openpyxl import load_workbook

from src.api.main import save_task_results_to_file

EXPECTED_HEADER = ["type", "platform", "post_id", "parent_id", "author", "content",
                   "published_at", "likes", "retweets", "replies", "tags"]


def _sample_stats():
    """模拟 pipeline 返回的 stats（含完整 crawled_posts：帖子 + 评论行）"""
    return {
        "status": "success",
        "analyze": False,
        "total_posts": 2,
        "total_keywords": 0,
        "platforms": [
            {"name": "twitter", "post_count": 1},
            {"name": "reddit", "post_count": 1},
        ],
        "crawled_posts": [
            {
                "platform": "twitter", "post_id": "t1", "parent_id": "", "author": "alice",
                "content": "hello world", "published_at": "2026-07-01T08:00:00",
                "likes": 10, "retweets": 2, "replies": 1, "tags": "a,b", "type": "post",
            },
            {
                "platform": "twitter", "post_id": "t1c1", "parent_id": "t1", "author": "carol",
                "content": "nice tweet", "published_at": "2026-07-01T08:05:00",
                "likes": 0, "retweets": 0, "replies": 0, "tags": "a,b", "type": "comment",
            },
            {
                "platform": "reddit", "post_id": "r1", "parent_id": "", "author": "bob",
                "content": "foo bar", "published_at": "2026-07-01T09:00:00",
                "likes": 5, "retweets": 0, "replies": 3, "tags": "", "type": "post",
            },
        ],
    }


async def test_save_crawl_only_csv(tmp_path):
    """CSV 完整写入帖子+评论明细，含 type/parent_id 列"""
    path = await save_task_results_to_file(
        save_path=str(tmp_path), task_name="测试任务",
        stats=_sample_stats(), results=None, file_format="csv",
    )
    assert path is not None
    assert path.endswith(".csv")
    assert os.path.exists(path)

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    # 首行表头 + 3 条数据（2 帖子 + 1 评论）
    assert rows[0] == EXPECTED_HEADER
    assert len(rows) == 4
    # 第一行：帖子
    assert rows[1][0] == "帖子" and rows[1][1] == "twitter" and rows[1][4] == "alice"
    # 第二行：评论，parent_id 指向所属帖子
    assert rows[2][0] == "评论" and rows[2][3] == "t1" and rows[2][5] == "nice tweet"
    # 第三行：reddit 帖子
    assert rows[3][0] == "帖子" and rows[3][1] == "reddit" and rows[3][5] == "foo bar"


async def test_save_crawl_only_xlsx(tmp_path):
    """xlsx 完整写入帖子+评论明细"""
    path = await save_task_results_to_file(
        save_path=str(tmp_path), task_name="测试任务",
        stats=_sample_stats(), results=None, file_format="xlsx",
    )
    assert path is not None
    assert path.endswith(".xlsx")

    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header[0] == "type" and "content" in header and "parent_id" in header
    # 表头 + 3 行数据
    assert ws.max_row == 4


async def test_save_uses_sampled_posts_fallback(tmp_path):
    """无 crawled_posts 时回退使用 sampled_posts"""
    stats = {
        "status": "success",
        "sampled_posts": [
            {"platform": "twitter", "post_id": "t1", "author": "alice",
             "content": "hi", "published_at": "2026-07-01T08:00:00",
             "metrics": {"likes": 3, "retweets": 1, "replies": 0}, "tags": []},
        ],
    }
    path = await save_task_results_to_file(
        save_path=str(tmp_path), task_name="t", stats=stats, file_format="csv",
    )
    assert path is not None
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    assert len(rows) == 2          # 表头 + 1 条
    assert rows[0] == EXPECTED_HEADER
    # likes 列现位于索引 7（type,platform,post_id,parent_id,author,content,published_at,likes,...）
    assert rows[1][7] == "3"       # metrics.likes 展平到 likes 列
    assert rows[1][0] == "帖子"     # 无 type 字段默认按帖子


async def test_save_no_path_returns_none():
    """未指定保存路径时返回 None（不落盘）"""
    path = await save_task_results_to_file(
        save_path="", task_name="t", stats=_sample_stats(), file_format="csv",
    )
    assert path is None
