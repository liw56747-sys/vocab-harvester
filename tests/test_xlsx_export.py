"""v1.5.5: XLSX 导出功能测试"""

import base64
import io
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook


@pytest.fixture
def client(monkeypatch, tmp_path):
    monkeypatch.setenv("VOCAB_DATA_DIR", str(tmp_path))
    from src.api.main import app
    return TestClient(app)


# ── 纯函数测试（无需数据库，CI 安全） ──

def test_generate_xlsx_bytes_basic():
    """测试 _generate_xlsx_bytes 生成基础 XLSX 文件"""
    from src.api.main import _generate_xlsx_bytes

    rows = [
        {"word": "test", "count": 1, "status": "approved"},
        {"word": "demo", "count": 2, "status": "pending"},
    ]
    fields = ["word", "count", "status"]
    sheet_title = "测试表"

    xlsx_bytes = _generate_xlsx_bytes(rows, fields, sheet_title)
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 0

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.title == sheet_title[:31]

    header_row = [cell.value for cell in ws[1]]
    assert header_row == fields

    assert ws.cell(2, 1).value == "test"
    assert ws.cell(2, 2).value == 1
    assert ws.cell(2, 3).value == "approved"
    assert ws.cell(3, 1).value == "demo"
    assert ws.cell(3, 2).value == 2
    assert ws.cell(3, 3).value == "pending"


def test_generate_xlsx_bytes_header_style():
    """测试表头样式：粗体 + 浅色底色"""
    from src.api.main import _generate_xlsx_bytes

    rows = [{"col1": "val1"}]
    fields = ["col1"]
    xlsx_bytes = _generate_xlsx_bytes(rows, fields, "样式测试")

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    header_cell = ws.cell(1, 1)
    assert header_cell.font.bold is True
    assert header_cell.fill.start_color.rgb is not None
    assert header_cell.alignment.vertical == "center"


def test_generate_xlsx_bytes_freeze_panes():
    """测试冻结首行"""
    from src.api.main import _generate_xlsx_bytes

    rows = [{"col1": "val1"}]
    fields = ["col1"]
    xlsx_bytes = _generate_xlsx_bytes(rows, fields, "冻结测试")

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.freeze_panes == "A2"


def test_generate_xlsx_bytes_column_width():
    """测试自适应列宽"""
    from src.api.main import _generate_xlsx_bytes

    rows = [
        {"short": "a", "long": "x" * 50},
    ]
    fields = ["short", "long"]
    xlsx_bytes = _generate_xlsx_bytes(rows, fields, "列宽测试")

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    from openpyxl.utils import get_column_letter
    short_width = ws.column_dimensions[get_column_letter(1)].width
    long_width = ws.column_dimensions[get_column_letter(2)].width

    assert short_width >= 8
    assert long_width >= 8
    assert long_width > short_width
    assert long_width <= 60


def test_generate_xlsx_bytes_boolean_handling():
    """测试布尔值转字符串"""
    from src.api.main import _generate_xlsx_bytes

    rows = [
        {"flag": True, "name": "test1"},
        {"flag": False, "name": "test2"},
    ]
    fields = ["flag", "name"]
    xlsx_bytes = _generate_xlsx_bytes(rows, fields, "布尔测试")

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    assert ws.cell(2, 1).value == "true"
    assert ws.cell(3, 1).value == "false"


def test_generate_xlsx_bytes_long_content_truncation():
    """测试超长内容截断（防止单元格膨胀）"""
    from src.api.main import _generate_xlsx_bytes

    long_text = "x" * 40000
    rows = [{"content": long_text}]
    fields = ["content"]
    xlsx_bytes = _generate_xlsx_bytes(rows, fields, "截断测试")

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    cell_value = ws.cell(2, 1).value
    assert len(cell_value) <= 32001
    assert cell_value.endswith("…")


# ── HTTP 端点测试（无需数据库初始化） ──

def test_download_xlsx_endpoint(client):
    """测试 /api/download-xlsx 端点"""
    from src.api.main import _generate_xlsx_bytes
    rows = [{"word": "test"}]
    fields = ["word"]
    xlsx_bytes = _generate_xlsx_bytes(rows, fields, "下载测试")
    xlsx_b64 = base64.b64encode(xlsx_bytes).decode("ascii")

    response = client.post(
        "/api/download-xlsx",
        json={"xlsx_data": xlsx_b64, "filename": "test.xlsx"}
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "test.xlsx" in response.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert ws.cell(1, 1).value == "word"
    assert ws.cell(2, 1).value == "test"


def test_download_xlsx_invalid_base64(client):
    """测试 /api/download-xlsx 对无效 base64 的处理"""
    response = client.post(
        "/api/download-xlsx",
        json={"xlsx_data": "invalid-base64!!!", "filename": "test.xlsx"}
    )

    assert response.status_code == 400
    assert "无效的 base64 数据" in response.json()["detail"]


# ── 集成测试（需要数据库，CI 跳过） ──

pytestmark_vocab = pytest.mark.integration


@pytest.mark.integration
def test_vocabulary_export_xlsx(monkeypatch, tmp_path):
    """测试词库导出 XLSX 格式（需要数据库初始化）"""
    import asyncio
    monkeypatch.setenv("VOCAB_DATA_DIR", str(tmp_path))
    from src.common.database import init_db
    db_path = tmp_path / "vocab.db"
    asyncio.run(init_db(db_path))

    from src.vocabulary.manager import VocabManager
    from src.common.models import VocabEntry, VocabStatus
    manager = VocabManager()

    entry1 = VocabEntry(word="test_word", category="test", status=VocabStatus.APPROVED)
    entry2 = VocabEntry(word="demo_word", category="demo", status=VocabStatus.PENDING)
    asyncio.run(manager.storage.upsert(entry1))
    asyncio.run(manager.storage.upsert(entry2))

    from src.api.main import app
    client = TestClient(app)
    response = client.get("/api/vocabulary/export?format=xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "vocabulary_export.xlsx" in response.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert ws.max_row >= 3


@pytest.mark.integration
def test_vocabulary_export_xlsx_with_status_filter(monkeypatch, tmp_path):
    """测试词库导出 XLSX 格式（带状态筛选）"""
    import asyncio
    monkeypatch.setenv("VOCAB_DATA_DIR", str(tmp_path))
    from src.common.database import init_db
    db_path = tmp_path / "vocab.db"
    asyncio.run(init_db(db_path))

    from src.vocabulary.manager import VocabManager
    from src.common.models import VocabEntry, VocabStatus
    manager = VocabManager()

    entry1 = VocabEntry(word="approved_word", category="test", status=VocabStatus.APPROVED)
    entry2 = VocabEntry(word="pending_word", category="test", status=VocabStatus.PENDING)
    asyncio.run(manager.storage.upsert(entry1))
    asyncio.run(manager.storage.upsert(entry2))

    from src.api.main import app
    client = TestClient(app)
    response = client.get("/api/vocabulary/export?format=xlsx&status=approved")

    assert response.status_code == 200

    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    word_col_idx = headers.index("word") + 1 if "word" in headers else 1

    words = [ws.cell(row, word_col_idx).value for row in range(2, ws.max_row + 1)]
    assert "approved_word" in words
    assert "pending_word" not in words


@pytest.mark.integration
def test_vocabulary_export_xlsx_empty(monkeypatch, tmp_path):
    """测试空词库导出 XLSX"""
    import asyncio
    monkeypatch.setenv("VOCAB_DATA_DIR", str(tmp_path))
    from src.common.database import init_db
    db_path = tmp_path / "empty_vocab.db"
    asyncio.run(init_db(db_path))

    from src.api.main import app
    client = TestClient(app)
    response = client.get("/api/vocabulary/export?format=xlsx")

    if response.status_code == 404:
        assert "词库为空" in response.json()["detail"]
    else:
        assert response.status_code == 200
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        assert ws.max_row == 1


@pytest.mark.integration
def test_batch_search_result_contains_xlsx_data():
    """测试 batch-search 结果包含 xlsx_data 字段（需要 Cookie）"""
    pytest.skip("需要 Cookie 配置，跳过集成测试")