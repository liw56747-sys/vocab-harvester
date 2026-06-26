"""
一键抓取 Twitter 用户推文 — 登录 + 抓取 + 导出 CSV/XLSX
用法: python fetch_xidong.py

首次运行需要输入 Twitter 账号信息，之后会保存 cookie 免登录。
"""

import asyncio
import csv
import traceback
from datetime import datetime
from pathlib import Path

# ── 配置 ─────────────────────────────────────────────
TARGET_USERNAME = "xidong6680370"  # 目标用户
TWEET_COUNT = 100                   # 抓取推文数量
COOKIE_PATH = Path(__file__).parent / ".twitter_cookies.json"
DATA_DIR = Path(__file__).parent / "data"

# CSV 字段
CSV_FIELDS = [
    "tweet_id", "author", "author_name", "content", "created_at",
    "likes", "retweets", "replies", "quotes", "url",
    "is_retweet", "is_reply",
]


async def ask_credentials():
    """交互式询问登录凭据"""
    print("\n" + "=" * 50)
    print("首次使用，需要登录 Twitter 账号")
    print("=" * 50)
    print("提示: 用户名填你的 Twitter 登录名（不是中文昵称）")
    print("      通常是 @ 后面那串英文/数字\n")

    username = input("Twitter 用户名 (如 elonmusk): ").strip()
    email = input("邮箱地址: ").strip()
    password = input("密码: ").strip()
    proxy = input("代理地址 (国内必填, 如 http://127.0.0.1:7890，留空跳过): ").strip()

    if not all([username, email, password]):
        print("[!] 所有字段都必须填写")
        return None

    return {
        "username": username,
        "email": email,
        "password": password,
        "proxy": proxy if proxy else None,
    }


async def do_login(client, username: str, email: str, password: str):
    """执行登录"""
    print(f"\n[*] 正在登录 @{username} ...")

    try:
        await client.login(
            auth_info_1=username,
            auth_info_2=email,
            password=password,
        )
    except Exception as e:
        print(f"\n[!] 登录失败!")
        print(f"    异常类型: {type(e).__name__}")
        print(f"    str(e):   '{str(e)}'")
        print(f"    repr(e):  {repr(e)}")
        if hasattr(e, 'args') and e.args:
            print(f"    args:     {e.args}")
        if hasattr(e, 'response'):
            resp = e.response
            print(f"    HTTP状态: {getattr(resp, 'status_code', 'N/A')}")
            try:
                print(f"    响应体:   {resp.text[:500]}")
            except Exception:
                pass
        traceback.print_exc()
        return False

    # save_cookies 是同步方法
    client.save_cookies(str(COOKIE_PATH))
    print(f"[+] 登录成功！Cookie 已保存")
    return True


async def load_cookies(client):
    """尝试加载已有 cookie"""
    if not COOKIE_PATH.exists():
        return False
    try:
        client.load_cookies(str(COOKIE_PATH))
        return True
    except Exception as e:
        print(f"[!] Cookie 加载失败: {e}")
        return False


async def fetch_tweets(client, username: str, count: int):
    """抓取指定用户的推文"""
    print(f"\n[*] 正在获取 @{username} 的推文 (目标 {count} 条)...")

    try:
        user = await client.get_user_by_screen_name(username)
    except Exception as e:
        print(f"[!] 用户 @{username} 获取失败: {type(e).__name__}: {e}")
        traceback.print_exc()
        return []

    print(f"[+] 找到用户: {getattr(user, 'name', '?')} (@{username})")

    tweets_data = []
    fetched = 0

    try:
        tweets = await user.get_tweets("Tweets", count=min(count, 20))

        while tweets and fetched < count:
            for tweet in tweets:
                if fetched >= count:
                    break

                tweet_url = f"https://x.com/{username}/status/{tweet.id}"
                tweets_data.append({
                    "tweet_id": str(tweet.id),
                    "author": f"@{username}",
                    "author_name": getattr(user, "name", username),
                    "content": tweet.text or "",
                    "created_at": str(tweet.created_at) if tweet.created_at else "",
                    "likes": getattr(tweet, "favorite_count", 0) or 0,
                    "retweets": getattr(tweet, "retweet_count", 0) or 0,
                    "replies": getattr(tweet, "reply_count", 0) or 0,
                    "quotes": getattr(tweet, "quote_count", 0) or 0,
                    "url": tweet_url,
                    "is_retweet": getattr(tweet, "retweeted_tweet", None) is not None,
                    "is_reply": getattr(tweet, "in_reply_to", None) is not None,
                })
                fetched += 1

            # 翻页
            if fetched < count:
                try:
                    tweets = await tweets.next()
                except Exception:
                    break

            print(f"    已抓取 {fetched} 条...")

    except Exception as e:
        print(f"[!] 抓取过程中出错: {type(e).__name__}: {e}")
        traceback.print_exc()

    return tweets_data


def save_csv(tweets: list[dict], username: str) -> str:
    """保存为 CSV"""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"tweets_{username}_{timestamp}.csv"
    filepath = DATA_DIR / filename

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        writer.writerows(tweets)

    print(f"[+] CSV 已保存: {filepath}")
    return str(filepath)


def save_xlsx(tweets: list[dict], username: str) -> str:
    """保存为 XLSX（带格式）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[!] openpyxl 未安装，跳过 XLSX 导出 (pip install openpyxl)")
        return ""

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"tweets_{username}_{timestamp}.xlsx"
    filepath = DATA_DIR / filename

    wb = Workbook()
    ws = wb.active
    ws.title = f"@{username}"

    # 写表头
    headers = ["推文ID", "作者", "昵称", "内容", "发布时间",
               "点赞", "转推", "回复", "引用", "链接",
               "是转推", "是回复"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1DA1F2")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 写数据
    field_map = ["tweet_id", "author", "author_name", "content", "created_at",
                 "likes", "retweets", "replies", "quotes", "url",
                 "is_retweet", "is_reply"]
    for row_idx, tweet in enumerate(tweets, 2):
        for col_idx, field in enumerate(field_map, 1):
            val = tweet.get(field, "")
            if isinstance(val, bool):
                val = "是" if val else "否"
            ws.cell(row=row_idx, column=col_idx, value=val)

    # 冻结首行、自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(str(filepath))
    print(f"[+] XLSX 已保存: {filepath}")
    return str(filepath)


async def main():
    import json

    print("=" * 60)
    print(f"  Twitter 推文抓取工具")
    print(f"  目标: @{TARGET_USERNAME}")
    print(f"  数量: {TWEET_COUNT} 条")
    print("=" * 60)

    from twikit import Client

    # 读取已保存的代理配置
    config_path = Path(__file__).parent / ".twitter_config.json"
    proxy = None
    if config_path.exists():
        try:
            cfg = json.loads(config_path.read_text(encoding="utf-8"))
            proxy = cfg.get("proxy") or None
        except Exception:
            pass

    # 1. 尝试加载 cookie
    client = Client("en-US", proxy=proxy)
    loaded = await load_cookies(client)
    if loaded:
        print(f"[+] 已加载保存的登录信息 (代理: {proxy or '无'})")
    else:
        # 需要登录
        creds = await ask_credentials()
        if not creds:
            input("\n按回车退出...")
            return

        # 用新代理重新创建 client
        proxy = creds["proxy"]
        client = Client("en-US", proxy=proxy)

        ok = await do_login(client, creds["username"], creds["email"], creds["password"])
        if not ok:
            print("\n[!] 登录未成功，请重试")
            input("\n按回车退出...")
            return

        # 保存代理配置
        config_path.write_text(json.dumps({"proxy": proxy or ""}, indent=2), encoding="utf-8")
        print(f"[+] 代理配置已保存: {proxy or '无'}")

    # 2. 抓取推文
    tweets = await fetch_tweets(client, TARGET_USERNAME, TWEET_COUNT)

    if not tweets:
        print("\n[!] 未抓取到任何推文，可能原因:")
        print("    - 该用户不存在或已被封禁")
        print("    - Cookie 已过期，需要重新登录")
        print("    - 网络问题（需要能访问 x.com）")
        if COOKIE_PATH.exists():
            COOKIE_PATH.unlink()
            print("    已清除旧 cookie，下次运行会重新登录")
        input("\n按回车退出...")
        return

    # 3. 导出
    print(f"\n[+] 成功抓取 {len(tweets)} 条推文")
    save_csv(tweets, TARGET_USERNAME)
    save_xlsx(tweets, TARGET_USERNAME)

    print("\n[完成] 文件保存在 data/ 目录下")
    input("\n按回车退出...")


if __name__ == "__main__":
    asyncio.run(main())
